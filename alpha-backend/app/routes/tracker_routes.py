from __future__ import annotations

import logging
from datetime import datetime
from typing import List, Optional

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.responses import StreamingResponse
from sqlmodel import Session, select

from app.core.config import settings
from app.db.tracker_db import get_tracker_session
from app.deps.auth import require_roles
from app.models.tracker.models import (
    TrackerApplication,
    TrackerApplicationAD,
    TrackerCandidate,
    TrackerCandidateAD,
    TrackerFollowUp,
    TrackerFollowUpAD,
    TrackerJobOpening,
    TrackerJobOpeningAD,
    TrackerOption,
    TrackerOptionAD,
)
from app.models.user import User
from app.schemas.tracker import (
    TrackerApplicationCreate,
    TrackerApplicationRead,
    TrackerApplicationUpdate,
    TrackerCandidateCreate,
    TrackerCandidateRead,
    TrackerCandidateRowRead,
    TrackerCandidateUpdate,
    TrackerFollowUpCreate,
    TrackerFollowUpRead,
    TrackerFollowUpUpdate,
    TrackerJobOpeningCreate,
    TrackerJobOpeningRead,
    TrackerJobOpeningUpdate,
    TrackerOptionCreate,
    TrackerOptionRead,
    TrackerOptionUpdate,
)
from app.utils.redis_cache import get_redis_cache
from app.core.config import is_followup_email_reminder_enabled
from app.services.followup_reminder_service import render_followup_email_html, send_followup_reminder_email

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/api/tracker", tags=["Candidate Tracker"])

_TRACKER_CACHE_NAMESPACE = "tracker"


def _tracker_cache_key(team: str, name: str, *parts: str) -> str:
    safe_parts = [p.replace(" ", "").replace("\n", "") for p in parts if p is not None]
    return "|".join([team, name, *safe_parts])


def _tracker_cache_get(team: str, name: str, *parts: str):
    try:
        return get_redis_cache().get(_tracker_cache_key(team, name, *parts), namespace=_TRACKER_CACHE_NAMESPACE)
    except Exception:
        return None


def _tracker_cache_set(team: str, name: str, value, ttl_seconds: int, *parts: str) -> None:
    try:
        get_redis_cache().set(
            _tracker_cache_key(team, name, *parts),
            value,
            ttl_seconds=ttl_seconds,
            namespace=_TRACKER_CACHE_NAMESPACE,
        )
    except Exception:
        return None


def _tracker_cache_bust() -> None:
    # Keep invalidation simple and safe: any write clears tracker cache only.
    try:
        get_redis_cache().clear_namespace(_TRACKER_CACHE_NAMESPACE)
    except Exception:
        return None


def tracker_session(session: Session = Depends(get_tracker_session)) -> Session:
    # IMPORTANT: this dependency runs before route handlers. Keep the feature flag and
    # TRACKER_DB_URL checks here so disabled tracker endpoints don't crash due to missing DB.
    _require_tracker_enabled()
    return session


def _require_tracker_enabled():
    if not settings.ENABLE_CANDIDATE_TRACKER:
        raise HTTPException(status_code=404, detail="Candidate Tracker is disabled")
    if not settings.TRACKER_DB_URL:
        raise HTTPException(status_code=500, detail="TRACKER_DB_URL not configured")


def _touch_updated_at(obj):
    if hasattr(obj, "updated_at"):
        obj.updated_at = datetime.utcnow()


# EVP has manager-level access across the app.
AllowedTrackerReadUser = Depends(require_roles("admin", "manager", "evp", "user", "recruiter"))
# Write access is restricted to managers/admins only (recruiters/users are read-only).
AllowedTrackerWriteUser = Depends(require_roles("manager"))
# Follow-ups: recruiters/users are allowed to create/update/delete.
AllowedTrackerFollowUpWriteUser = Depends(require_roles("admin", "manager", "evp", "user", "recruiter"))
AllowedTrackerJobWriteUser = Depends(require_roles("manager"))
AllowedTrackerOptionWriteUser = Depends(require_roles("manager"))
AllowedTrackerAdminUser = Depends(require_roles("admin"))
AllowedTrackerManagerUser = Depends(require_roles("manager"))

_RECRUITER_ROLES = frozenset({"recruiter", "user"})


def _norm_team(v: str) -> str:
    return "".join(ch for ch in (v or "").strip().lower() if ch.isalnum())


def _resolve_tracker_team(user: User, team: Optional[str]) -> str:
    """
    Resolve which tracker table set to use.

    - Non-admin/EVP: forced by user.team_location.
    - Admin/EVP: can override via ?team=dubai|abudhabi.
    """
    is_admin_like = user.role in {"admin", "evp"}
    if is_admin_like and team:
        t = _norm_team(team)
        if t in {"dubai", "dxb"}:
            return "dubai"
        if t in {"abudhabi", "abudabi", "abudh"}:
            return "abudhabi"
        raise HTTPException(status_code=400, detail="Invalid team; use dubai|abudhabi")

    tl = _norm_team(getattr(user, "team_location", None) or "")
    if tl in {"abudhabi", "abudabi", "abudh"}:
        return "abudhabi"
    # default to dubai to preserve existing behavior/data
    return "dubai"


def _tracker_models(team: str):
    if team == "abudhabi":
        return {
            "JobOpening": TrackerJobOpeningAD,
            "Candidate": TrackerCandidateAD,
            "Application": TrackerApplicationAD,
            "FollowUp": TrackerFollowUpAD,
            "Option": TrackerOptionAD,
        }
    return {
        "JobOpening": TrackerJobOpening,
        "Candidate": TrackerCandidate,
        "Application": TrackerApplication,
        "FollowUp": TrackerFollowUp,
        "Option": TrackerOption,
    }


def _reject_job_status_option_for_candidate_roles(user: User, kind: str | None) -> None:
    """Recruiters (`user` | `recruiter`) must not edit requirement (job) status picklists."""
    if kind == "job_status" and user.role in _RECRUITER_ROLES:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Not allowed to modify requirement status options",
        )


@router.post("/admin/bootstrap-abudhabi", status_code=200)
def bootstrap_abudhabi_from_existing_tables(
    session: Session = Depends(tracker_session),
    _: User = AllowedTrackerAdminUser,
):
    """
    One-time bootstrap helper (non-destructive):
    Copy rows from the existing tracker tables into the Abu Dhabi *_ad tables.

    - Does NOT delete or modify Dubai tables.
    - Keeps the same IDs so relationships (candidate_id, job_opening_id, etc.) remain valid.
    - Safe to run multiple times (idempotent best-effort).
    """
    from sqlalchemy import text

    bind = session.get_bind()
    url = str(getattr(bind, "url", ""))
    is_sqlite = url.startswith("sqlite")

    # Include all current tracker tables (even if some endpoints are unused today).
    # Copy order: referenced tables first.
    table_pairs = [
        ("tracker_option", "tracker_option_ad"),
        ("tracker_job_opening", "tracker_job_opening_ad"),
        ("tracker_candidate", "tracker_candidate_ad"),
        ("tracker_application", "tracker_application_ad"),
        ("tracker_follow_up", "tracker_follow_up_ad"),
    ]

    inserted: dict[str, int] = {}

    with bind.begin() as conn:
        for src, dst in table_pairs:
            try:
                if is_sqlite:
                    # SQLite supports INSERT OR IGNORE for PK conflicts.
                    res = conn.execute(text(f'INSERT OR IGNORE INTO "{dst}" SELECT * FROM "{src}"'))
                else:
                    # Postgres: use ON CONFLICT DO NOTHING on the primary key (id).
                    res = conn.execute(text(f'INSERT INTO "{dst}" SELECT * FROM "{src}" ON CONFLICT (id) DO NOTHING'))
                inserted[dst] = int(getattr(res, "rowcount", 0) or 0)
            except Exception:
                # Best-effort: if table doesn't exist or schemas diverge, skip.
                inserted[dst] = inserted.get(dst, 0)

    _tracker_cache_bust()
    return {"success": True, "inserted": inserted}


@router.post("/admin/import/selections-xlsx", status_code=200)
async def import_selections_xlsx(
    team: str = Query(..., description="dubai|abudhabi"),
    file: UploadFile = File(...),
    session: Session = Depends(tracker_session),
    user: User = AllowedTrackerAdminUser,
):
    """
    Admin import for Selections & Joinings from the exported candidates.xlsx format.

    - Imports into the selected team's tables (Dubai or Abu Dhabi).
    - Does not touch Manager Settings/options; you can manage those manually.
    """
    from datetime import date
    from io import BytesIO

    from openpyxl import load_workbook

    team_norm = _norm_team(team)
    if team_norm not in {"dubai", "dxb", "abudhabi", "abudabi", "abudh"}:
        raise HTTPException(status_code=400, detail="Invalid team; use dubai|abudhabi")
    team_resolved = "abudhabi" if team_norm in {"abudhabi", "abudabi", "abudh"} else "dubai"

    models = _tracker_models(team_resolved)
    Candidate = models["Candidate"]
    Application = models["Application"]

    raw = await file.read()
    if not raw:
        raise HTTPException(status_code=400, detail="Empty file")

    try:
        wb = load_workbook(filename=BytesIO(raw), data_only=True)
        ws = wb.active
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid Excel file")

    # Expect the same header as export_candidates_xlsx
    header_row = [str(c.value).strip() if c.value is not None else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
    header_map = {h.lower(): i for i, h in enumerate(header_row)}
    required = ["candidate name"]
    for r in required:
        if r not in header_map:
            raise HTTPException(status_code=400, detail=f"Missing required column: {r}")

    def _cell(row, name: str) -> str:
        idx = header_map.get(name.lower())
        if idx is None:
            return ""
        v = row[idx].value
        return (str(v).strip() if v is not None else "")

    def _parse_date(v: str | None) -> date | None:
        if not v:
            return None
        try:
            # If openpyxl already gave us a python date/datetime, keep it.
            if hasattr(v, "year") and hasattr(v, "month") and hasattr(v, "day"):
                return date(int(v.year), int(v.month), int(v.day))  # type: ignore[arg-type]
            s = str(v).strip()
            if not s:
                return None
            return date.fromisoformat(s[:10])
        except Exception:
            return None

    inserted_candidates = 0
    inserted_applications = 0
    skipped = 0

    for row in ws.iter_rows(min_row=2):
        name = _cell(row, "Candidate Name")
        if not name:
            skipped += 1
            continue

        cand = session.exec(
            select(Candidate).where(Candidate.name == name).where(Candidate.is_deleted == False)  # noqa: E712
        ).first()
        if not cand:
            cand = Candidate(name=name)
            session.add(cand)
            session.commit()
            session.refresh(cand)
            inserted_candidates += 1

        applied_date = _parse_date(_cell(row, "Date"))
        app = Application(
            candidate_id=cand.id,
            job_opening_id=None,
            applied_date=applied_date,
            position=_cell(row, "Position") or None,
            client=_cell(row, "Client") or None,
            status=_cell(row, "Status") or "MRF Pending",
            recruiter=_cell(row, "Recruiter") or None,
            account_manager=_cell(row, "Account Manager") or None,
            recruitment_manager=_cell(row, "Recruitment Manager") or None,
            comment=_cell(row, "Comment") or None,
            created_by_username=getattr(user, "username", None),
            created_by_role=getattr(user, "role", None),
            created_by_team_location=("Abu Dhabi" if team_resolved == "abudhabi" else "Dubai"),
        )
        session.add(app)
        session.commit()
        inserted_applications += 1

    _tracker_cache_bust()
    return {
        "success": True,
        "team": team_resolved,
        "inserted_candidates": inserted_candidates,
        "inserted_applications": inserted_applications,
        "skipped_rows": skipped,
    }


@router.post("/admin/clear-team", status_code=200)
def clear_tracker_team_data(
    team: str = Query(..., description="dubai|abudhabi"),
    confirm: str = Query(..., description="must be CLEAR to proceed"),
    session: Session = Depends(tracker_session),
    _: User = AllowedTrackerAdminUser,
):
    """
    Admin-only bulk clear for a team's tracker tables.

    This is intended for the migration path where you import old data into Abu Dhabi
    and then want to clear the Dubai dataset without touching Abu Dhabi (or vice versa).
    """
    from sqlalchemy import text

    if confirm.strip().upper() != "CLEAR":
        raise HTTPException(status_code=400, detail="Missing confirm=CLEAR")

    team_norm = _norm_team(team)
    if team_norm not in {"dubai", "dxb", "abudhabi", "abudabi", "abudh"}:
        raise HTTPException(status_code=400, detail="Invalid team; use dubai|abudhabi")

    is_ad = team_norm in {"abudhabi", "abudabi", "abudh"}

    # IMPORTANT: delete order matters due to foreign keys.
    # We clear all tracker tables for that team, including options.
    tables = (
        [
            "tracker_application_ad",
            "tracker_candidate_ad",
            "tracker_job_opening_ad",
            "tracker_follow_up_ad",
            "tracker_option_ad",
        ]
        if is_ad
        else [
            "tracker_application",
            "tracker_candidate",
            "tracker_job_opening",
            "tracker_follow_up",
            "tracker_option",
        ]
    )

    bind = session.get_bind()
    deleted: dict[str, int] = {}
    with bind.begin() as conn:
        for t in tables:
            try:
                res = conn.execute(text(f'DELETE FROM "{t}"'))
                deleted[t] = int(getattr(res, "rowcount", 0) or 0)
            except Exception:
                # Best-effort: if table doesn't exist in this deployment, skip.
                deleted[t] = deleted.get(t, 0)

    _tracker_cache_bust()
    return {"success": True, "team": ("abudhabi" if is_ad else "dubai"), "deleted": deleted}


@router.get("/job-openings", response_model=List[TrackerJobOpeningRead])
def list_job_openings(
    status_filter: Optional[str] = None,
    team: Optional[str] = Query(default=None),
    session: Session = Depends(tracker_session),
    user: User = AllowedTrackerReadUser,
):
    team_resolved = _resolve_tracker_team(user, team)
    cache_hit = _tracker_cache_get(
        team_resolved,
        "job_openings",
        f"status={status_filter or ''}",
    )
    if cache_hit is not None:
        return cache_hit
    JobOpening = _tracker_models(team_resolved)["JobOpening"]
    q = select(JobOpening).where(JobOpening.is_deleted == False)  # noqa: E712
    if status_filter:
        q = q.where(JobOpening.status == status_filter)
    rows = session.exec(q.order_by(JobOpening.created_at.desc())).all()
    out = [
        TrackerJobOpeningRead(
            id=r.id,
            title=r.title,
            department=r.department,
            client=getattr(r, "client", None),
            status=r.status,
            hiring_manager=r.hiring_manager,
            recruitment_manager=getattr(r, "recruitment_manager", None),
            req_date=r.req_date,
            submission_date=getattr(r, "submission_date", None),
            cvs_submitted_count=getattr(r, "cvs_submitted_count", None),
            comments=getattr(r, "comments", None),
            created_at=r.created_at,
            updated_at=r.updated_at,
        )
        for r in rows
    ]
    _tracker_cache_set(team_resolved, "job_openings", [o.model_dump() for o in out], 60, f"status={status_filter or ''}")
    return out


@router.post("/job-openings", response_model=TrackerJobOpeningRead, status_code=201)
def create_job_opening(
    data: TrackerJobOpeningCreate,
    team: Optional[str] = Query(default=None),
    session: Session = Depends(tracker_session),
    user: User = AllowedTrackerJobWriteUser,
):
    team_resolved = _resolve_tracker_team(user, team)
    JobOpening = _tracker_models(team_resolved)["JobOpening"]
    row = JobOpening(
        title=data.title,
        department=data.department,
        client=getattr(data, "client", None),
        status=data.status,
        hiring_manager=data.hiring_manager,
        recruitment_manager=getattr(data, "recruitment_manager", None),
        req_date=data.req_date,
        submission_date=getattr(data, "submission_date", None),
        cvs_submitted_count=getattr(data, "cvs_submitted_count", None),
        comments=getattr(data, "comments", None),
    )
    session.add(row)
    session.commit()
    session.refresh(row)
    _tracker_cache_bust()
    return TrackerJobOpeningRead(
        id=row.id,
        title=row.title,
        department=row.department,
        client=getattr(row, "client", None),
        status=row.status,
        hiring_manager=row.hiring_manager,
        recruitment_manager=getattr(row, "recruitment_manager", None),
        req_date=row.req_date,
        submission_date=getattr(row, "submission_date", None),
        cvs_submitted_count=getattr(row, "cvs_submitted_count", None),
        comments=getattr(row, "comments", None),
        created_at=row.created_at,
        updated_at=row.updated_at,
    )


@router.patch("/job-openings/{job_opening_id}", response_model=TrackerJobOpeningRead)
def update_job_opening(
    job_opening_id: str,
    data: TrackerJobOpeningUpdate,
    team: Optional[str] = Query(default=None),
    session: Session = Depends(tracker_session),
    user: User = AllowedTrackerJobWriteUser,
):
    team_resolved = _resolve_tracker_team(user, team)
    JobOpening = _tracker_models(team_resolved)["JobOpening"]
    row = session.exec(select(JobOpening).where(JobOpening.id == job_opening_id)).first()
    if not row or row.is_deleted:
        raise HTTPException(status_code=404, detail="Job opening not found")
    for k, v in data.model_dump(exclude_unset=True).items():
        setattr(row, k, v)
    _touch_updated_at(row)
    session.add(row)
    session.commit()
    session.refresh(row)
    _tracker_cache_bust()
    return TrackerJobOpeningRead(
        id=row.id,
        title=row.title,
        department=row.department,
        client=getattr(row, "client", None),
        status=row.status,
        hiring_manager=row.hiring_manager,
        recruitment_manager=getattr(row, "recruitment_manager", None),
        req_date=row.req_date,
        submission_date=getattr(row, "submission_date", None),
        cvs_submitted_count=getattr(row, "cvs_submitted_count", None),
        comments=getattr(row, "comments", None),
        created_at=row.created_at,
        updated_at=row.updated_at,
    )


@router.get("/options", response_model=List[TrackerOptionRead])
def list_options(
    kind: str,
    include_deleted: bool = False,
    team: Optional[str] = Query(default=None),
    session: Session = Depends(tracker_session),
    user: User = AllowedTrackerReadUser,
):
    team_resolved = _resolve_tracker_team(user, team)
    cache_hit = _tracker_cache_get(
        team_resolved,
        "options",
        f"kind={kind}",
        f"include_deleted={include_deleted}",
    )
    if cache_hit is not None:
        return cache_hit
    Option = _tracker_models(team_resolved)["Option"]
    q = select(Option).where(Option.kind == kind)
    if not include_deleted:
        q = q.where(Option.is_deleted == False)  # noqa: E712
    rows = session.exec(q.order_by(Option.value.asc())).all()
    out = [
        TrackerOptionRead(
            id=r.id,
            kind=r.kind,
            value=r.value,
            email=getattr(r, "email", None),
            email_enabled=bool(getattr(r, "email_enabled", True)),
            is_deleted=r.is_deleted,
            created_at=r.created_at,
            updated_at=r.updated_at,
        )
        for r in rows
    ]
    _tracker_cache_set(
        team_resolved,
        "options",
        [o.model_dump() for o in out],
        120,
        f"kind={kind}",
        f"include_deleted={include_deleted}",
    )
    return out


@router.post("/options", response_model=TrackerOptionRead, status_code=201)
def create_option(
    data: TrackerOptionCreate,
    team: Optional[str] = Query(default=None),
    session: Session = Depends(tracker_session),
    user: User = AllowedTrackerOptionWriteUser,
):
    team_resolved = _resolve_tracker_team(user, team)
    Option = _tracker_models(team_resolved)["Option"]
    kind = data.kind.strip()
    _reject_job_status_option_for_candidate_roles(user, kind)
    value = data.value.strip()
    if not kind or not value:
        raise HTTPException(status_code=400, detail="kind and value are required")

    existing = session.exec(
        select(Option).where(Option.kind == kind).where(Option.value == value).where(Option.is_deleted == False)  # noqa: E712
    ).first()
    if existing:
        return TrackerOptionRead(
            id=existing.id,
            kind=existing.kind,
            value=existing.value,
            email=getattr(existing, "email", None),
            email_enabled=bool(getattr(existing, "email_enabled", True)),
            is_deleted=existing.is_deleted,
            created_at=existing.created_at,
            updated_at=existing.updated_at,
        )

    row = Option(
        kind=kind,
        value=value,
        email=(data.email.strip() if isinstance(getattr(data, "email", None), str) and data.email.strip() else None),
        email_enabled=bool(getattr(data, "email_enabled", True)),
    )
    session.add(row)
    session.commit()
    session.refresh(row)
    _tracker_cache_bust()
    return TrackerOptionRead(
        id=row.id,
        kind=row.kind,
        value=row.value,
        email=getattr(row, "email", None),
        email_enabled=bool(getattr(row, "email_enabled", True)),
        is_deleted=row.is_deleted,
        created_at=row.created_at,
        updated_at=row.updated_at,
    )


@router.delete("/options/purge-inactive", status_code=200)
def purge_inactive_options(
    kind: str,
    team: Optional[str] = Query(default=None),
    session: Session = Depends(tracker_session),
    user: User = AllowedTrackerOptionWriteUser,
):
    """Hard-delete inactive (is_deleted=true) options for a kind.

    This only affects tracker_option rows and never deletes any requirement/joinings/follow-ups records.
    """
    if not kind or not kind.strip():
        raise HTTPException(status_code=400, detail="kind is required")
    kind = kind.strip()
    team_resolved = _resolve_tracker_team(user, team)
    Option = _tracker_models(team_resolved)["Option"]
    rows = session.exec(
        select(Option).where(Option.kind == kind).where(Option.is_deleted == True)  # noqa: E712
    ).all()
    n = 0
    for r in rows:
        session.delete(r)
        n += 1
    session.commit()
    if n:
        _tracker_cache_bust()
    return {"kind": kind, "purged": n}


@router.patch("/options/{option_id}", response_model=TrackerOptionRead)
def update_option(
    option_id: str,
    data: TrackerOptionUpdate,
    team: Optional[str] = Query(default=None),
    session: Session = Depends(tracker_session),
    user: User = AllowedTrackerOptionWriteUser,
):
    team_resolved = _resolve_tracker_team(user, team)
    Option = _tracker_models(team_resolved)["Option"]
    row = session.exec(select(Option).where(Option.id == option_id)).first()
    if not row or row.is_deleted:
        raise HTTPException(status_code=404, detail="Option not found")
    _reject_job_status_option_for_candidate_roles(user, row.kind)
    any_change = False
    if data.value is not None:
        row.value = data.value.strip()
        any_change = True
    if getattr(data, "email", None) is not None:
        v = (data.email or "").strip()
        setattr(row, "email", v or None)
        any_change = True
    if getattr(data, "email_enabled", None) is not None:
        setattr(row, "email_enabled", bool(data.email_enabled))
        any_change = True
    if not any_change:
        raise HTTPException(status_code=400, detail="No fields to update")
    _touch_updated_at(row)
    session.add(row)
    session.commit()
    session.refresh(row)
    _tracker_cache_bust()
    return TrackerOptionRead(
        id=row.id,
        kind=row.kind,
        value=row.value,
        email=getattr(row, "email", None),
        email_enabled=bool(getattr(row, "email_enabled", True)),
        is_deleted=row.is_deleted,
        created_at=row.created_at,
        updated_at=row.updated_at,
    )


@router.delete("/options/{option_id}", status_code=204)
def delete_option(
    option_id: str,
    team: Optional[str] = Query(default=None),
    session: Session = Depends(tracker_session),
    user: User = AllowedTrackerOptionWriteUser,
):
    team_resolved = _resolve_tracker_team(user, team)
    Option = _tracker_models(team_resolved)["Option"]
    row = session.exec(select(Option).where(Option.id == option_id)).first()
    if not row or row.is_deleted:
        raise HTTPException(status_code=404, detail="Option not found")
    _reject_job_status_option_for_candidate_roles(user, row.kind)
    row.is_deleted = True
    _touch_updated_at(row)
    session.add(row)
    session.commit()
    _tracker_cache_bust()
    return None


@router.delete("/job-openings/{job_opening_id}", status_code=204)
def delete_job_opening(
    job_opening_id: str,
    team: Optional[str] = Query(default=None),
    session: Session = Depends(tracker_session),
    user: User = AllowedTrackerJobWriteUser,
):
    team_resolved = _resolve_tracker_team(user, team)
    JobOpening = _tracker_models(team_resolved)["JobOpening"]
    row = session.exec(select(JobOpening).where(JobOpening.id == job_opening_id)).first()
    if not row or row.is_deleted:
        raise HTTPException(status_code=404, detail="Job opening not found")
    row.is_deleted = True
    _touch_updated_at(row)
    session.add(row)
    session.commit()
    _tracker_cache_bust()
    return None


@router.get("/candidates", response_model=List[TrackerCandidateRead])
def list_candidates(
    team: Optional[str] = Query(default=None),
    session: Session = Depends(tracker_session),
    user: User = AllowedTrackerReadUser,
):
    team_resolved = _resolve_tracker_team(user, team)
    Candidate = _tracker_models(team_resolved)["Candidate"]
    rows = session.exec(
        select(Candidate).where(Candidate.is_deleted == False).order_by(Candidate.created_at.desc())  # noqa: E712
    ).all()
    return [
        TrackerCandidateRead(
            id=r.id,
            name=r.name,
            email=r.email,
            phone=r.phone,
            current_company=r.current_company,
            experience_years=r.experience_years,
            created_at=r.created_at,
            updated_at=r.updated_at,
        )
        for r in rows
    ]


@router.post("/candidates", response_model=TrackerCandidateRead, status_code=201)
def create_candidate(
    data: TrackerCandidateCreate,
    team: Optional[str] = Query(default=None),
    session: Session = Depends(tracker_session),
    user: User = AllowedTrackerWriteUser,
):
    team_resolved = _resolve_tracker_team(user, team)
    Candidate = _tracker_models(team_resolved)["Candidate"]
    row = Candidate(
        name=data.name,
        email=data.email,
        phone=data.phone,
        current_company=data.current_company,
        experience_years=data.experience_years,
    )
    session.add(row)
    session.commit()
    session.refresh(row)
    _tracker_cache_bust()
    return TrackerCandidateRead(
        id=row.id,
        name=row.name,
        email=row.email,
        phone=row.phone,
        current_company=row.current_company,
        experience_years=row.experience_years,
        created_at=row.created_at,
        updated_at=row.updated_at,
    )


@router.patch("/candidates/{candidate_id}", response_model=TrackerCandidateRead)
def update_candidate(
    candidate_id: str,
    data: TrackerCandidateUpdate,
    team: Optional[str] = Query(default=None),
    session: Session = Depends(tracker_session),
    user: User = AllowedTrackerWriteUser,
):
    team_resolved = _resolve_tracker_team(user, team)
    Candidate = _tracker_models(team_resolved)["Candidate"]
    row = session.exec(select(Candidate).where(Candidate.id == candidate_id)).first()
    if not row or row.is_deleted:
        raise HTTPException(status_code=404, detail="Candidate not found")
    for k, v in data.model_dump(exclude_unset=True).items():
        setattr(row, k, v)
    _touch_updated_at(row)
    session.add(row)
    session.commit()
    session.refresh(row)
    _tracker_cache_bust()
    return TrackerCandidateRead(
        id=row.id,
        name=row.name,
        email=row.email,
        phone=row.phone,
        current_company=row.current_company,
        experience_years=row.experience_years,
        created_at=row.created_at,
        updated_at=row.updated_at,
    )


@router.delete("/candidates/{candidate_id}", status_code=204)
def delete_candidate(
    candidate_id: str,
    team: Optional[str] = Query(default=None),
    session: Session = Depends(tracker_session),
    user: User = AllowedTrackerWriteUser,
):
    team_resolved = _resolve_tracker_team(user, team)
    Candidate = _tracker_models(team_resolved)["Candidate"]
    row = session.exec(select(Candidate).where(Candidate.id == candidate_id)).first()
    if not row or row.is_deleted:
        raise HTTPException(status_code=404, detail="Candidate not found")
    row.is_deleted = True
    _touch_updated_at(row)
    session.add(row)
    session.commit()
    _tracker_cache_bust()
    return None


@router.post("/applications", response_model=TrackerApplicationRead, status_code=201)
def create_application(
    data: TrackerApplicationCreate,
    team: Optional[str] = Query(default=None),
    session: Session = Depends(tracker_session),
    user: User = AllowedTrackerWriteUser,
):
    team_resolved = _resolve_tracker_team(user, team)
    models = _tracker_models(team_resolved)
    Candidate = models["Candidate"]
    JobOpening = models["JobOpening"]
    Application = models["Application"]
    # Validate candidate exists
    cand = session.exec(select(Candidate).where(Candidate.id == data.candidate_id)).first()
    if not cand or cand.is_deleted:
        raise HTTPException(status_code=400, detail="Invalid candidate_id")
    job_opening_id = (data.job_opening_id or "").strip() or None
    # job_opening_id is optional — only validate if provided
    if job_opening_id:
        job = session.exec(select(JobOpening).where(JobOpening.id == job_opening_id)).first()
        if not job or job.is_deleted:
            raise HTTPException(status_code=400, detail="Invalid job_opening_id")

    row = Application(
        candidate_id=data.candidate_id,
        job_opening_id=job_opening_id,
        applied_date=data.applied_date,
        position=data.position,
        client=data.client,
        status=data.status,
        recruiter=data.recruiter,
        account_manager=data.account_manager,
        recruitment_manager=getattr(data, "recruitment_manager", None),
        comment=data.comment,
        created_by_username=getattr(user, "username", None),
        created_by_role=getattr(user, "role", None),
        created_by_team_location=getattr(user, "team_location", None),
    )
    session.add(row)
    session.commit()
    session.refresh(row)
    _tracker_cache_bust()
    return TrackerApplicationRead(
        id=row.id,
        candidate_id=row.candidate_id,
        job_opening_id=row.job_opening_id,
        applied_date=row.applied_date,
        position=row.position,
        client=row.client,
        status=row.status,
        recruiter=row.recruiter,
        account_manager=row.account_manager,
        recruitment_manager=getattr(row, "recruitment_manager", None),
        comment=row.comment,
        created_at=row.created_at,
        updated_at=row.updated_at,
    )


@router.get("/applications", response_model=List[TrackerApplicationRead])
def list_applications(
    candidate_id: Optional[str] = None,
    job_opening_id: Optional[str] = None,
    team: Optional[str] = Query(default=None),
    session: Session = Depends(tracker_session),
    user: User = AllowedTrackerReadUser,
):
    team_resolved = _resolve_tracker_team(user, team)
    Application = _tracker_models(team_resolved)["Application"]
    q = select(Application)
    if candidate_id:
        q = q.where(Application.candidate_id == candidate_id)
    if job_opening_id:
        q = q.where(Application.job_opening_id == job_opening_id)
    rows = session.exec(q.order_by(Application.created_at.desc())).all()
    return [
        TrackerApplicationRead(
            id=r.id,
            candidate_id=r.candidate_id,
            job_opening_id=r.job_opening_id,
            applied_date=r.applied_date,
            position=r.position,
            client=r.client,
            status=r.status,
            recruiter=r.recruiter,
            account_manager=r.account_manager,
            recruitment_manager=getattr(r, "recruitment_manager", None),
            comment=r.comment,
            created_at=r.created_at,
            updated_at=r.updated_at,
        )
        for r in rows
    ]


@router.get("/candidate-rows", response_model=List[TrackerCandidateRowRead])
def list_candidate_rows(
    team: Optional[str] = Query(default=None),
    session: Session = Depends(tracker_session),
    user: User = AllowedTrackerReadUser,
):
    """Convenience endpoint for UI: returns candidate + latest application in one call."""
    team_resolved = _resolve_tracker_team(user, team)
    cache_hit = _tracker_cache_get(
        team_resolved,
        "candidate_rows",
        f"role={user.role}",
        f"team_location={_norm_team(getattr(user, 'team_location', None) or '')}",
    )
    if cache_hit is not None:
        return cache_hit
    models = _tracker_models(team_resolved)
    Candidate = models["Candidate"]
    Application = models["Application"]
    candidates = session.exec(
        select(Candidate).where(Candidate.is_deleted == False).order_by(Candidate.created_at.desc())  # noqa: E712
    ).all()
    apps = session.exec(select(Application).order_by(Application.created_at.desc())).all()

    # Row-level access safety:
    # Even though we split tables by team, older deployments may have mixed-location rows
    # in the Dubai tables from before the split. For non-admin/EVP users, enforce a strict
    # filter by the immutable creator team location (fallback: editable location only if needed).
    if user.role not in {"admin", "evp"}:
        team_loc = _norm_team(getattr(user, "team_location", None) or "")
        if team_loc:
            apps = [
                a
                for a in apps
                if _norm_team(
                    (getattr(a, "created_by_team_location", None) or "")
                )
                == team_loc
            ]
        else:
            apps = []

    latest_app_by_cand: dict[str, Application] = {}
    for a in apps:
        if a.candidate_id not in latest_app_by_cand:
            latest_app_by_cand[a.candidate_id] = a

    out: list[TrackerCandidateRowRead] = []
    for c in candidates:
        a = latest_app_by_cand.get(c.id)
        # For restricted users, hide candidates with no visible application in their team.
        if user.role not in {"admin", "evp"} and not a:
            continue
        out.append(
            TrackerCandidateRowRead(
                candidate=TrackerCandidateRead(
                    id=c.id,
                    name=c.name,
                    email=c.email,
                    phone=c.phone,
                    current_company=c.current_company,
                    experience_years=c.experience_years,
                    created_at=c.created_at,
                    updated_at=c.updated_at,
                ),
                application=(
                    TrackerApplicationRead(
                        id=a.id,
                        candidate_id=a.candidate_id,
                        job_opening_id=a.job_opening_id,
                        applied_date=a.applied_date,
                        position=a.position,
                        client=a.client,
                        status=a.status,
                        recruiter=a.recruiter,
                        account_manager=a.account_manager,
                        recruitment_manager=getattr(a, "recruitment_manager", None),
                        comment=a.comment,
                        created_at=a.created_at,
                        updated_at=a.updated_at,
                    )
                    if a
                    else None
                ),
            )
        )
    # Keep TTL short because this is the heaviest endpoint.
    _tracker_cache_set(
        team_resolved,
        "candidate_rows",
        [o.model_dump() for o in out],
        30,
        f"role={user.role}",
        f"team_location={_norm_team(getattr(user, 'team_location', None) or '')}",
    )
    return out


@router.patch("/applications/{application_id}", response_model=TrackerApplicationRead)
def update_application(
    application_id: str,
    data: TrackerApplicationUpdate,
    team: Optional[str] = Query(default=None),
    session: Session = Depends(tracker_session),
    user: User = AllowedTrackerWriteUser,
):
    team_resolved = _resolve_tracker_team(user, team)
    Application = _tracker_models(team_resolved)["Application"]
    row = session.exec(select(Application).where(Application.id == application_id)).first()
    if not row:
        raise HTTPException(status_code=404, detail="Application not found")
    for k, v in data.model_dump(exclude_unset=True).items():
        setattr(row, k, v)
    _touch_updated_at(row)
    session.add(row)
    session.commit()
    session.refresh(row)
    _tracker_cache_bust()
    return TrackerApplicationRead(
        id=row.id,
        candidate_id=row.candidate_id,
        job_opening_id=row.job_opening_id,
        applied_date=row.applied_date,
        position=row.position,
        client=row.client,
        status=row.status,
        recruiter=row.recruiter,
        account_manager=row.account_manager,
        recruitment_manager=getattr(row, "recruitment_manager", None),
        comment=row.comment,
        created_at=row.created_at,
        updated_at=row.updated_at,
    )


@router.get("/follow-ups", response_model=List[TrackerFollowUpRead])
def list_followups(
    team: Optional[str] = Query(default=None),
    session: Session = Depends(tracker_session),
    user: User = AllowedTrackerReadUser,
):
    team_resolved = _resolve_tracker_team(user, team)
    cache_hit = _tracker_cache_get(team_resolved, "followups")
    if cache_hit is not None:
        return cache_hit
    FollowUp = _tracker_models(team_resolved)["FollowUp"]
    rows = session.exec(
        select(FollowUp)
        .where(FollowUp.is_deleted == False)  # noqa: E712
        .order_by(FollowUp.next_follow_up_date.asc(), FollowUp.created_at.desc())
    ).all()
    out = [
        TrackerFollowUpRead(
            id=r.id,
            client_name=r.client_name,
            position=r.position,
            recruiter_name=r.recruiter_name,
            account_manager=getattr(r, "account_manager", None),
            recruitment_manager=getattr(r, "recruitment_manager", None),
            cv_submitted_date=r.cv_submitted_date,
            current_stage=r.current_stage,
            last_follow_up_date=r.last_follow_up_date,
            next_follow_up_date=r.next_follow_up_date,
            interview_date=r.interview_date,
            client_feedback=r.client_feedback,
            interview_feedback=r.interview_feedback,
            remarks=r.remarks,
            reminder_last_sent_at=getattr(r, "reminder_last_sent_at", None),
            created_at=r.created_at,
            updated_at=r.updated_at,
        )
        for r in rows
    ]
    _tracker_cache_set(team_resolved, "followups", [o.model_dump() for o in out], 60)
    return out


@router.post("/follow-ups", response_model=TrackerFollowUpRead, status_code=201)
def create_followup(
    data: TrackerFollowUpCreate,
    team: Optional[str] = Query(default=None),
    session: Session = Depends(tracker_session),
    user: User = AllowedTrackerFollowUpWriteUser,
):
    team_resolved = _resolve_tracker_team(user, team)
    FollowUp = _tracker_models(team_resolved)["FollowUp"]
    row = FollowUp(
        client_name=data.client_name.strip(),
        position=(data.position.strip() if data.position else None),
        recruiter_name=(data.recruiter_name.strip() if data.recruiter_name else None),
        account_manager=(data.account_manager.strip() if getattr(data, "account_manager", None) else None),
        recruitment_manager=(data.recruitment_manager.strip() if getattr(data, "recruitment_manager", None) else None),
        cv_submitted_date=data.cv_submitted_date,
        current_stage=data.current_stage.strip(),
        last_follow_up_date=data.last_follow_up_date,
        next_follow_up_date=data.next_follow_up_date,
        interview_date=data.interview_date,
        client_feedback=data.client_feedback,
        interview_feedback=data.interview_feedback,
        remarks=data.remarks,
    )
    session.add(row)
    session.commit()
    session.refresh(row)
    _tracker_cache_bust()
    return TrackerFollowUpRead(
        id=row.id,
        client_name=row.client_name,
        position=row.position,
        recruiter_name=row.recruiter_name,
        account_manager=getattr(row, "account_manager", None),
        recruitment_manager=getattr(row, "recruitment_manager", None),
        cv_submitted_date=row.cv_submitted_date,
        current_stage=row.current_stage,
        last_follow_up_date=row.last_follow_up_date,
        next_follow_up_date=row.next_follow_up_date,
        interview_date=row.interview_date,
        client_feedback=row.client_feedback,
        interview_feedback=row.interview_feedback,
        remarks=row.remarks,
        reminder_last_sent_at=getattr(row, "reminder_last_sent_at", None),
        created_at=row.created_at,
        updated_at=row.updated_at,
    )


@router.get("/manager-settings/followup-email", status_code=200)
def get_followup_email_settings(
    team: Optional[str] = Query(default=None),
    session: Session = Depends(tracker_session),
    user: User = AllowedTrackerManagerUser,
):
    team_resolved = _resolve_tracker_team(user, team)
    Option = _tracker_models(team_resolved)["Option"]

    to_row = session.exec(
        select(Option)
        .where(Option.kind == "followup_email_to")
        .where(Option.is_deleted == False)  # noqa: E712
        .order_by(Option.updated_at.desc())
    ).first()
    cc_row = session.exec(
        select(Option)
        .where(Option.kind == "followup_email_cc")
        .where(Option.is_deleted == False)  # noqa: E712
        .order_by(Option.updated_at.desc())
    ).first()

    return {
        "team": team_resolved,
        "to": (to_row.value if to_row else ""),
        "cc": (cc_row.value if cc_row else ""),
        "enabled": is_followup_email_reminder_enabled(),
    }


@router.put("/manager-settings/followup-email", status_code=200)
def set_followup_email_settings(
    to: str = Query(..., description="comma-separated list of To emails"),
    cc: Optional[str] = Query(default=None, description="comma-separated list of CC emails (optional)"),
    team: Optional[str] = Query(default=None),
    session: Session = Depends(tracker_session),
    user: User = AllowedTrackerManagerUser,
):
    team_resolved = _resolve_tracker_team(user, team)
    Option = _tracker_models(team_resolved)["Option"]

    def upsert(kind: str, value: str):
        existing = session.exec(
            select(Option)
            .where(Option.kind == kind)
            .where(Option.is_deleted == False)  # noqa: E712
            .order_by(Option.updated_at.desc())
        ).first()
        if existing:
            existing.value = value.strip()
            _touch_updated_at(existing)
            session.add(existing)
            session.commit()
            session.refresh(existing)
            return existing
        row = Option(kind=kind, value=value.strip())
        session.add(row)
        session.commit()
        session.refresh(row)
        return row

    upsert("followup_email_to", to)
    # CC can be empty; store empty string for clarity.
    upsert("followup_email_cc", (cc or ""))
    _tracker_cache_bust()
    return {"success": True, "team": team_resolved, "to": to, "cc": (cc or ""), "enabled": is_followup_email_reminder_enabled()}


@router.post("/follow-ups/{followup_id}/send-reminder", status_code=200)
async def send_followup_reminder_now(
    followup_id: str,
    team: Optional[str] = Query(default=None),
    to: Optional[str] = Query(default=None, description="override To list (optional)"),
    cc: Optional[str] = Query(default=None, description="override CC list (optional)"),
    session: Session = Depends(tracker_session),
    user: User = AllowedTrackerManagerUser,
):
    """
    Manager-triggered "Send now" reminder for a single follow-up row.
    Uses Manager Settings To/CC by default; can be overridden per request.
    """
    if not is_followup_email_reminder_enabled():
        raise HTTPException(status_code=400, detail="Follow-up email reminders are disabled (SEND_EMAIL_REMINDER_FOLLOWUP!=true)")

    team_resolved = _resolve_tracker_team(user, team)
    models = _tracker_models(team_resolved)
    FollowUp = models["FollowUp"]
    Option = models["Option"]

    row = session.exec(select(FollowUp).where(FollowUp.id == followup_id)).first()
    if not row or row.is_deleted:
        raise HTTPException(status_code=404, detail="Follow-up not found")

    # Never send reminders when the current stage indicates the position is closed.
    stage_raw = str(getattr(row, "current_stage", "") or "").strip().lower()
    stage_norm = " ".join(stage_raw.split())
    if stage_norm in {"positions closed", "position closed"}:
        raise HTTPException(status_code=400, detail="Reminder not sent because Current Stage is Positions Closed.")

    def get_opt(kind: str) -> str:
        r = session.exec(
            select(Option)
            .where(Option.kind == kind)
            .where(Option.is_deleted == False)  # noqa: E712
            .order_by(Option.updated_at.desc())
        ).first()
        return (r.value if r else "").strip()

    def lookup_person_email(kind: str, name: Optional[str]) -> tuple[str, bool]:
        if not name or not str(name).strip():
            return "", True
        nm = str(name).strip()
        r = session.exec(
            select(Option)
            .where(Option.kind == kind)
            .where(Option.value == nm)
            .where(Option.is_deleted == False)  # noqa: E712
            .order_by(Option.updated_at.desc())
        ).first()
        if not r:
            return "", True
        email = (getattr(r, "email", None) or "").strip()
        enabled = bool(getattr(r, "email_enabled", True))
        return email, enabled

    def _split_emails(v: str) -> list[str]:
        parts = []
        for raw in str(v or "").replace(";", ",").split(","):
            e = raw.strip()
            if e:
                parts.append(e)
        out: list[str] = []
        seen = set()
        for e in parts:
            k = e.lower()
            if k in seen:
                continue
            seen.add(k)
            out.append(e)
        return out

    def _join_emails(items: list[str]) -> str:
        return ", ".join([e.strip() for e in items if str(e or "").strip()])

    to_override = (to or "").strip()
    cc_override = (cc or "").strip()

    # To: Manager Settings (or override). CC: default CC + selected people (or override).
    # Account Manager has an ON/OFF toggle controlling whether reminders should be sent for their rows.
    _am_email, am_enabled = lookup_person_email("account_manager", getattr(row, "account_manager", None))
    if not am_enabled:
        raise HTTPException(status_code=400, detail="Account Manager reminders are OFF for this person.")

    to_final = to_override or get_opt("followup_email_to")
    if not to_final:
        raise HTTPException(status_code=400, detail="Missing To emails. Set it in Manager Settings first.")

    if cc_override:
        cc_final = cc_override
    else:
        base_cc = _split_emails(get_opt("followup_email_cc"))
        rec_email, _rec_enabled = lookup_person_email("recruiter", getattr(row, "recruiter_name", None))
        am_email, _am_enabled2 = lookup_person_email("account_manager", getattr(row, "account_manager", None))
        rm_email, _rm_enabled = lookup_person_email("recruitment_manager", getattr(row, "recruitment_manager", None))
        cc_items = base_cc + _split_emails(rec_email) + _split_emails(am_email) + _split_emails(rm_email)
        cc_final = _join_emails(cc_items)

    payload = TrackerFollowUpRead(
        id=row.id,
        client_name=row.client_name,
        position=row.position,
        recruiter_name=row.recruiter_name,
        account_manager=getattr(row, "account_manager", None),
        recruitment_manager=getattr(row, "recruitment_manager", None),
        cv_submitted_date=row.cv_submitted_date,
        current_stage=row.current_stage,
        last_follow_up_date=row.last_follow_up_date,
        next_follow_up_date=row.next_follow_up_date,
        client_feedback=row.client_feedback,
        interview_feedback=row.interview_feedback,
        remarks=row.remarks,
        reminder_last_sent_at=getattr(row, "reminder_last_sent_at", None),
        created_at=row.created_at,
        updated_at=row.updated_at,
    ).model_dump()

    subject = f"Follow-up Reminder: {row.client_name}" + (f" – {row.position}" if row.position else "")
    html = render_followup_email_html(payload)
    ok = await send_followup_reminder_email(to_emails=to_final, cc_emails=cc_final or None, subject=subject, body_html=html)
    if not ok:
        raise HTTPException(status_code=500, detail="Failed to send reminder email")

    row.reminder_last_sent_at = datetime.utcnow()
    _touch_updated_at(row)
    session.add(row)
    session.commit()
    _tracker_cache_bust()
    return {"success": True, "team": team_resolved, "id": row.id, "to": to_final, "cc": cc_final}


@router.patch("/follow-ups/{followup_id}", response_model=TrackerFollowUpRead)
def update_followup(
    followup_id: str,
    data: TrackerFollowUpUpdate,
    team: Optional[str] = Query(default=None),
    session: Session = Depends(tracker_session),
    user: User = AllowedTrackerFollowUpWriteUser,
):
    team_resolved = _resolve_tracker_team(user, team)
    FollowUp = _tracker_models(team_resolved)["FollowUp"]
    row = session.exec(select(FollowUp).where(FollowUp.id == followup_id)).first()
    if not row or row.is_deleted:
        raise HTTPException(status_code=404, detail="Follow-up not found")
    for k, v in data.model_dump(exclude_unset=True).items():
        if isinstance(v, str):
            v = v.strip()
        setattr(row, k, v)
    _touch_updated_at(row)
    session.add(row)
    session.commit()
    session.refresh(row)
    _tracker_cache_bust()
    return TrackerFollowUpRead(
        id=row.id,
        client_name=row.client_name,
        position=row.position,
        recruiter_name=row.recruiter_name,
        account_manager=getattr(row, "account_manager", None),
        recruitment_manager=getattr(row, "recruitment_manager", None),
        cv_submitted_date=row.cv_submitted_date,
        current_stage=row.current_stage,
        last_follow_up_date=row.last_follow_up_date,
        next_follow_up_date=row.next_follow_up_date,
        interview_date=row.interview_date,
        client_feedback=row.client_feedback,
        interview_feedback=row.interview_feedback,
        remarks=row.remarks,
        reminder_last_sent_at=getattr(row, "reminder_last_sent_at", None),
        created_at=row.created_at,
        updated_at=row.updated_at,
    )


@router.delete("/follow-ups/{followup_id}", status_code=204)
def delete_followup(
    followup_id: str,
    team: Optional[str] = Query(default=None),
    session: Session = Depends(tracker_session),
    user: User = AllowedTrackerFollowUpWriteUser,
):
    team_resolved = _resolve_tracker_team(user, team)
    FollowUp = _tracker_models(team_resolved)["FollowUp"]
    row = session.exec(select(FollowUp).where(FollowUp.id == followup_id)).first()
    if not row or row.is_deleted:
        raise HTTPException(status_code=404, detail="Follow-up not found")
    row.is_deleted = True
    _touch_updated_at(row)
    session.add(row)
    session.commit()
    _tracker_cache_bust()
    return None


@router.get("/export/job-openings.xlsx")
def export_job_openings_xlsx(
    team: Optional[str] = Query(default=None),
    session: Session = Depends(tracker_session),
    user: User = AllowedTrackerReadUser,
):
    from io import BytesIO
    from openpyxl import Workbook

    team_resolved = _resolve_tracker_team(user, team)
    JobOpening = _tracker_models(team_resolved)["JobOpening"]
    rows = session.exec(
        select(JobOpening)
        .where(JobOpening.is_deleted == False)  # noqa: E712
        .order_by(JobOpening.created_at.desc())
    ).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Requirement Status"
    ws.append(
        [
            "S.no",
            "Requirement",
            "Role",
            "Client",
            "Recruiter",
            "Recruiter Manager",
            "Account Manager",
            "Requirement Date",
            "Submission Date",
            "No of CVs Submitted",
            "Status",
            "Comments",
        ]
    )
    for i, r in enumerate(rows, start=1):
        ws.append(
            [
                i,
                getattr(r, "requirement", None) or "",
                r.title or "",
                getattr(r, "client", None) or "",
                r.hiring_manager or "",
                getattr(r, "recruitment_manager", None) or "",
                r.department or "",
                r.req_date.isoformat() if getattr(r, "req_date", None) else "",
                getattr(r, "submission_date", None).isoformat() if getattr(r, "submission_date", None) else "",
                getattr(r, "cvs_submitted_count", None) if getattr(r, "cvs_submitted_count", None) is not None else "",
                r.status or "",
                getattr(r, "comments", None) or "",
            ]
        )

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    headers = {"Content-Disposition": 'attachment; filename="job_openings.xlsx"'}
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )


@router.get("/export/candidates.xlsx")
def export_candidates_xlsx(
    team: Optional[str] = Query(default=None),
    session: Session = Depends(tracker_session),
    user: User = AllowedTrackerReadUser,
):
    from io import BytesIO
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Candidates"
    ws.append(
        [
            "Date",
            "Candidate Name",
            "Position",
            "Client",
            "Status",
            "Recruiter",
            "Recruitment Manager",
            "Account Manager",
            "Comment",
        ]
    )

    team_resolved = _resolve_tracker_team(user, team)
    models = _tracker_models(team_resolved)
    Candidate = models["Candidate"]
    Application = models["Application"]
    candidates_by_id = {
        c.id: c
        for c in session.exec(select(Candidate).where(Candidate.is_deleted == False)).all()  # noqa: E712
    }
    apps = session.exec(select(Application).order_by(Application.created_at.desc())).all()

    for a in apps:
        cand = candidates_by_id.get(a.candidate_id)
        if not cand:
            continue
        ws.append(
            [
                a.applied_date.isoformat() if a.applied_date else "",
                cand.name,
                a.position or "",
                a.client or "",
                a.status or "",
                a.recruiter or "",
                getattr(a, "recruitment_manager", None) or "",
                a.account_manager or "",
                a.comment or "",
            ]
        )

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    headers = {"Content-Disposition": 'attachment; filename="candidates.xlsx"'}
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )


@router.get("/export/follow-ups.xlsx")
def export_followups_xlsx(
    team: Optional[str] = Query(default=None),
    session: Session = Depends(tracker_session),
    user: User = AllowedTrackerReadUser,
):
    from io import BytesIO
    from openpyxl import Workbook

    team_resolved = _resolve_tracker_team(user, team)
    FollowUp = _tracker_models(team_resolved)["FollowUp"]
    rows = session.exec(
        select(FollowUp)
        .where(FollowUp.is_deleted == False)  # noqa: E712
        .order_by(FollowUp.next_follow_up_date.asc(), FollowUp.created_at.desc())
    ).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Follow-ups"
    ws.append(
        [
            "Client Name",
            "Position",
            "Recruiter Name",
            "Recruitment Manager",
            "CV Submitted Date",
            "Current Stage",
            "Follow Up Date",
            "Next Follow-up Date",
            "Interview Date",
            "Client Feedback",
            "Interview Feedback",
            "Remarks",
        ]
    )
    for r in rows:
        ws.append(
            [
                r.client_name or "",
                r.position or "",
                r.recruiter_name or "",
                getattr(r, "recruitment_manager", None) or "",
                r.cv_submitted_date.isoformat() if r.cv_submitted_date else "",
                r.current_stage or "",
                r.last_follow_up_date.isoformat() if r.last_follow_up_date else "",
                r.next_follow_up_date.isoformat() if r.next_follow_up_date else "",
                r.interview_date.isoformat() if r.interview_date else "",
                r.client_feedback or "",
                r.interview_feedback or "",
                r.remarks or "",
            ]
        )

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    headers = {"Content-Disposition": 'attachment; filename="follow_ups.xlsx"'}
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )

